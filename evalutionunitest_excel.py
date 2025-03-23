import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def generate_test_cases_excel(output_filename="qa_evaluation_test_cases.xlsx"):
    """
    Generate an Excel file with the test case documentation for the QA evaluation script.
    
    Args:
        output_filename (str): Name of the output Excel file
    """
    # Create test case data
    test_cases = [
        {
            "TEST CASE ID": "TC-001",
            "SECTION": "Metrics",
            "SUB-SECTION": "Calculation",
            "TEST CASE TITLE": "Verify metric calculation accuracy",
            "TEST DESCRIPTION": "Test that the calculate_metrics function returns accurate scores for identical, different, and partially similar texts",
            "PRECONDITIONS": "Python environment with required libraries installed",
            "TEST DATA": "Text pairs for comparison: identical, different, and partially similar",
            "TEST STEPS": "1. Call calculate_metrics with identical texts\n2. Call calculate_metrics with different texts\n3. Call calculate_metrics with partially similar texts",
            "EXPECTED RESULT": "1. identical_metrics: rouge_score=1.0, cosine_sim>0.99, bert_f1>0.99\n2. different_metrics: rouge_score<0.5, cosine_sim<0.8\n3. partial_metrics: 0.3<rouge<1.0, 0.7<cosine<1.0",
            "ACTUAL RESULT": "As expected",
            "STATUS": "PASS"
        },
        {
            "TEST CASE ID": "TC-002",
            "SECTION": "Grading",
            "SUB-SECTION": "Calculation",
            "TEST CASE TITLE": "Verify grade calculation logic",
            "TEST DESCRIPTION": "Test that the calculate_grade function correctly converts numerical scores to letter grades",
            "PRECONDITIONS": "Python environment with required libraries installed",
            "TEST DATA": "Test scores: 0.95, 0.85, 0.75, 0.65, 0.55",
            "TEST STEPS": "1. Call calculate_grade with each test score",
            "EXPECTED RESULT": "1. 0.95 -> \"A (Excellent)\"\n2. 0.85 -> \"B (Good)\"\n3. 0.75 -> \"C (Average)\"\n4. 0.65 -> \"D (Below Average)\"\n5. 0.55 -> \"F (Poor)\"",
            "ACTUAL RESULT": "As expected",
            "STATUS": "PASS"
        },
        {
            "TEST CASE ID": "TC-003",
            "SECTION": "Logging",
            "SUB-SECTION": "Functionality",
            "TEST CASE TITLE": "Verify log interaction functionality",
            "TEST DESCRIPTION": "Test that log_interaction properly writes event logs with all required fields",
            "PRECONDITIONS": "Python environment with required libraries installed, write access to log file",
            "TEST DATA": "Question, context, generated answer, reference answer, and metrics dictionary",
            "TEST STEPS": "1. Set up logger\n2. Call log_interaction with test data\n3. Read log file content",
            "EXPECTED RESULT": "Log file contains question text, context, and all metrics fields",
            "ACTUAL RESULT": "As expected",
            "STATUS": "PASS"
        },
        {
            "TEST CASE ID": "TC-004",
            "SECTION": "Data Loading",
            "SUB-SECTION": "File Handling",
            "TEST CASE TITLE": "Verify test case loading from Excel",
            "TEST DESCRIPTION": "Test that load_test_cases correctly loads data from Excel files",
            "PRECONDITIONS": "Python environment with pandas and required libraries installed, test Excel file",
            "TEST DATA": "Temporary Excel file with Question, Answer, and Context columns",
            "TEST STEPS": "1. Create temporary Excel file\n2. Call load_test_cases with file path\n3. Verify returned DataFrame",
            "EXPECTED RESULT": "DataFrame contains expected columns and rows",
            "ACTUAL RESULT": "As expected",
            "STATUS": "PASS"
        },
        {
            "TEST CASE ID": "TC-005",
            "SECTION": "API",
            "SUB-SECTION": "Integration",
            "TEST CASE TITLE": "Verify Mistral API interaction",
            "TEST DESCRIPTION": "Test that qa_pipeline correctly sends requests to Mistral API and processes responses",
            "PRECONDITIONS": "Python environment with required libraries installed, mock API response",
            "TEST DATA": "Mock response with sample answer",
            "TEST STEPS": "1. Mock requests.post to return success response\n2. Call qa_pipeline with question and context\n3. Verify returned answer and API call parameters",
            "EXPECTED RESULT": "Returns expected answer and calls API with correct parameters",
            "ACTUAL RESULT": "As expected",
            "STATUS": "PASS"
        },
        {
            "TEST CASE ID": "TC-006",
            "SECTION": "API",
            "SUB-SECTION": "Error Handling",
            "TEST CASE TITLE": "Verify API retry mechanism",
            "TEST DESCRIPTION": "Test that qa_pipeline correctly implements exponential backoff for rate limit errors",
            "PRECONDITIONS": "Python environment with required libraries installed, mock API responses for rate limit and success",
            "TEST DATA": "Mock responses for rate limit error and subsequent success",
            "TEST STEPS": "1. Mock requests.post to return rate limit error then success\n2. Mock time.sleep to avoid waiting\n3. Call qa_pipeline and verify behavior",
            "EXPECTED RESULT": "Retries after rate limit error and returns correct answer on second attempt",
            "ACTUAL RESULT": "As expected",
            "STATUS": "PASS"
        },
        {
            "TEST CASE ID": "TC-007",
            "SECTION": "Configuration",
            "SUB-SECTION": "Validation",
            "TEST CASE TITLE": "Verify model validation",
            "TEST DESCRIPTION": "Test that the script correctly validates the configured model against allowed models",
            "PRECONDITIONS": "Python environment with required libraries installed",
            "TEST DATA": "Invalid model name",
            "TEST STEPS": "1. Patch MISTRAL_MODEL with invalid value\n2. Attempt to import the module\n3. Verify ValueError is raised with appropriate message",
            "EXPECTED RESULT": "ValueError raised with \"Invalid model\" message",
            "ACTUAL RESULT": "As expected",
            "STATUS": "PASS"
        },
        {
            "TEST CASE ID": "TC-008",
            "SECTION": "End-to-End",
            "SUB-SECTION": "Integration",
            "TEST CASE TITLE": "Verify complete processing pipeline",
            "TEST DESCRIPTION": "Test the full test case processing workflow from loading to results",
            "PRECONDITIONS": "Python environment with required libraries, test files",
            "TEST DATA": "Sample test cases file with 2-3 entries",
            "TEST STEPS": "1. Setup temporary files\n2. Mock API responses\n3. Call process_test_cases\n4. Verify results file and metrics",
            "EXPECTED RESULT": "Results file contains expected entries and metrics",
            "ACTUAL RESULT": "Not executed",
            "STATUS": "N/A"
        }
    ]
    
    # Create DataFrame from test cases
    df = pd.DataFrame(test_cases)
    
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Add headers
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Add data
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Auto-adjust column widths
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=len(test_cases) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    # Set row heights
    for i in range(2, len(test_cases) + 2):
        ws.row_dimensions[i].height = 80  # Fixed height for data rows
    
    ws.row_dimensions[1].height = 40  # Header row height
    
    # Color status column conditionally
    for row_idx, status in enumerate(df['STATUS'].values, 2):
        cell = ws.cell(row=row_idx, column=len(headers))
        if status == 'PASS':
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif status == 'FAIL':
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif status == 'N/A':
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Save the workbook
    wb.save(output_filename)
    print(f"Excel file created successfully: {output_filename}")

if __name__ == "__main__":
    generate_test_cases_excel()