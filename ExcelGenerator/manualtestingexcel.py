import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_test_cases_excel(output_file="MedlinePlus_Test_Cases.xlsx"):
    """
    Generate an Excel file with test cases for the MedlinePlus Scraper and Vectorizer.
    
    Args:
        output_file: Name of the output Excel file
    """
    # Define the test cases data
    test_cases = [
        # Environment Tests
        ["TC-ENV-001", "Environment", "API Key", "Verify Mistral API Key availability", 
         "Check if the Mistral API key is properly loaded from environment variables", 
         ".env file should exist", "N/A", 
         "1. Call test_api_key() function", 
         "Function returns True and displays \"✓ Mistral API key is available\"", "", ""],
        
        ["TC-ENV-002", "Environment", "API Key", "Verify behavior with missing Mistral API key", 
         "Test system behavior when Mistral API key is not available", 
         "Remove MISTRAL_API_KEY from environment", "N/A", 
         "1. Temporarily rename .env file or remove API key\n2. Call test_api_key() function\n3. Restore original .env file", 
         "Function returns False and displays \"✗ Mistral API key is missing\"", "", ""],
        
        ["TC-ENV-003", "Environment", "Directory Creation", "Verify output directory creation", 
         "Check if the system creates the output directory if it doesn't exist", 
         "Output directory should not exist", "output_dir = \"test_medlineplus_output\"", 
         "1. Delete the output directory if it exists\n2. Initialize MedlinePlusScraper with the test output directory", 
         "Directory is created successfully with message \"Created output directory: test_medlineplus_output\"", "", ""],
        
        # Scraper Tests
        ["TC-SCR-001", "Scraper", "URL Retrieval", "Verify webpage retrieval with valid URL", 
         "Test if the scraper can retrieve content from a valid MedlinePlus URL", 
         "Internet connection available", "url = \"https://medlineplus.gov/ency/article/000147.htm\"", 
         "1. Call scraper.retrieve_webpage(url)", 
         "Function returns HTML content as string", "", ""],
        
        ["TC-SCR-002", "Scraper", "URL Retrieval", "Verify webpage retrieval with invalid URL", 
         "Test system behavior when an invalid URL is provided", 
         "Internet connection available", "url = \"https://medlineplus.gov/invalid/url.htm\"", 
         "1. Call scraper.retrieve_webpage(url)", 
         "Function returns None and prints error message", "", ""],
        
        ["TC-SCR-003", "Scraper", "URL Retrieval", "Verify webpage retrieval with connection error", 
         "Test system behavior when connection fails", 
         "Internet connection disrupted or timeout", "url = \"https://medlineplus.gov/ency/article/000147.htm\"", 
         "1. Disable internet connection\n2. Call scraper.retrieve_webpage(url)\n3. Restore internet connection", 
         "Function returns None and prints error message", "", ""],
        
        ["TC-SCR-004", "Scraper", "Content Parsing", "Verify article content parsing", 
         "Test if the scraper can extract article content correctly", 
         "Valid HTML content available", "html = content from valid MedlinePlus page", 
         "1. Retrieve HTML from a valid article URL\n2. Call scraper.parse_article_content(html)", 
         "Function returns dictionary with article title and sections", "", ""],
        
        ["TC-SCR-005", "Scraper", "Content Parsing", "Verify content parsing with invalid HTML", 
         "Test system behavior when HTML is malformed", 
         "Invalid HTML content", "html = \"<html><body>Incomplete HTML\"", 
         "1. Call scraper.parse_article_content(html)", 
         "Function returns error dictionary and prints error message", "", ""],
        
        ["TC-SCR-006", "Scraper", "File Operations", "Verify safe filename creation", 
         "Test if safe filenames are created from article titles", 
         "N/A", "title = \"Alzheimer's Disease/Dementia: A?<Test>\"", 
         "1. Call scraper.create_safe_filename(title)", 
         "Function returns filename with invalid characters removed, spaces replaced with underscores, and timestamp appended", "", ""],
        
        ["TC-SCR-007", "Scraper", "File Operations", "Verify file saving", 
         "Test if article content is saved properly to a file", 
         "Output directory exists", "content = {\"Title\": \"Test Disease\", \"Description\": \"Test content\"}\nurl = \"https://test.url\"", 
         "1. Call scraper.save_to_file(content, url)", 
         "Function returns filepath to saved file with content properly formatted", "", ""],
        
        ["TC-SCR-008", "Scraper", "Encyclopedia Links", "Verify article link extraction", 
         "Test if article links are correctly extracted for a letter", 
         "Internet connection available", "letter = \"Z\"", 
         "1. Call scraper.find_encyclopedia_articles(letter)", 
         "Function returns list of article URLs for the letter Z", "", ""],
        
        ["TC-SCR-009", "Scraper", "Encyclopedia Links", "Verify article link extraction with invalid letter", 
         "Test system behavior with invalid letter input", 
         "N/A", "letter = \"123\"", 
         "1. Call scraper.find_encyclopedia_articles(letter)", 
         "Function raises ValueError", "", ""],
        
        ["TC-SCR-010", "Scraper", "Main Process", "Verify end-to-end scraping process", 
         "Test the complete scraping process for a letter", 
         "Internet connection available\nOutput directory exists", "letter = \"Z\"\nlimit articles to 2", 
         "1. Call test_scraper(\"Z\", 2)", 
         "Function creates files for 2 articles from letter Z in the output directory", "", ""],
        
        # Vectorizer Tests
        ["TC-VEC-001", "Vectorizer", "File Combination", "Verify file combination", 
         "Test if multiple files are combined correctly", 
         "At least 2 test files exist in input directory", "input_dir = \"test_medlineplus_output\"", 
         "1. Create test files with known content\n2. Initialize vectorizer with test directory\n3. Call vectorizer.combine_files()", 
         "Function returns combined text containing content from all files with document separators", "", ""],
        
        ["TC-VEC-002", "Vectorizer", "File Combination", "Verify file combination with empty directory", 
         "Test system behavior when input directory is empty", 
         "Empty input directory", "input_dir = \"empty_test_dir\"", 
         "1. Create empty test directory\n2. Initialize vectorizer with empty directory\n3. Call vectorizer.combine_files()", 
         "Function returns empty string and reports 0 files combined", "", ""],
        
        ["TC-VEC-003", "Vectorizer", "Chunking", "Verify text chunking", 
         "Test if text is properly split into chunks", 
         "Combined text available", "text = output from combine_files()", 
         "1. Call vectorizer.create_chunks(text)", 
         "Function returns list of Document objects with appropriate metadata", "", ""],
        
        ["TC-VEC-004", "Vectorizer", "Chunking", "Verify chunking with short text", 
         "Test system behavior when text is shorter than chunk size", 
         "Short test text", "text = \"Short test text\"", 
         "1. Call vectorizer.create_chunks(text)", 
         "Function returns list with single Document containing the full text", "", ""],
        
        ["TC-VEC-005", "Vectorizer", "Vector Database", "Verify vector database creation", 
         "Test if vector database is created correctly", 
         "Document chunks available", "documents = output from create_chunks()", 
         "1. Call vectorizer.create_vector_db(documents)", 
         "Function creates ChromaDB database with correct number of entries", "", ""],
        
        ["TC-VEC-006", "Vectorizer", "Vector Database", "Verify vector database persistence", 
         "Test if vector database is properly persisted", 
         "Vector database has been created", "N/A", 
         "1. Create vector database\n2. Check if files exist in ./chroma_db directory", 
         "ChromaDB files exist in the specified directory", "", ""],
        
        ["TC-VEC-007", "Vectorizer", "RAG Pipeline", "Verify RAG pipeline initialization", 
         "Test if RAG pipeline is initialized correctly", 
         "Vector database exists\nMistral API key available", "N/A", 
         "1. Call vectorizer.initialize_rag_pipeline()", 
         "Function returns a RetrievalQA object without errors", "", ""],
        
        ["TC-VEC-008", "Vectorizer", "RAG Pipeline", "Verify RAG query processing", 
         "Test if RAG pipeline processes queries correctly", 
         "Vector database exists\nMistral API key available", "query = \"What are the symptoms of Alzheimer's disease?\"", 
         "1. Call vectorizer.query_with_rag(query)", 
         "Function returns response tuple containing answer and empty context string", "", ""],
        
        ["TC-VEC-009", "Vectorizer", "RAG Pipeline", "Verify RAG pipeline with missing API key", 
         "Test system behavior when Mistral API key is missing", 
         "Vector database exists\nMistral API key unavailable", "query = \"What is diabetes?\"", 
         "1. Temporarily remove API key\n2. Call vectorizer.query_with_rag(query)\n3. Restore API key", 
         "Function raises ValueError about missing API key", "", ""],
        
        ["TC-VEC-010", "Vectorizer", "Main Process", "Verify complete vectorization process", 
         "Test the end-to-end vectorization process", 
         "Test files exist in input directory\nMistral API key available", "input_dir = \"test_medlineplus_output\"", 
         "1. Call vectorizer.process()", 
         "Function completes all steps (combine, chunk, create vector DB) without errors", "", ""],
        
        # Integration Tests
        ["TC-INT-001", "Integration", "End-to-End", "Verify scraper to vectorizer workflow", 
         "Test the complete workflow from scraping to querying", 
         "Internet connection available\nMistral API key available", "letter = \"Z\"\nlimit = 2\nquery = \"What is Zinc deficiency?\"", 
         "1. Run test_scraper(\"Z\", 2)\n2. Initialize vectorizer with output directory\n3. Run vectorizer.process()\n4. Run vectorizer.query_with_rag(query)", 
         "Complete workflow executes without errors and returns relevant answer to query", "", ""],
        
        ["TC-INT-002", "Integration", "Error Handling", "Verify error handling in workflow", 
         "Test system resilience when errors occur in the workflow", 
         "Intentional error conditions", "N/A", 
         "1. Create test case with invalid directory\n2. Run full process and observe error handling", 
         "System gracefully handles errors and provides appropriate error messages", "", ""],
        
        # CLI Tests
        ["TC-CLI-001", "CLI", "User Interface", "Verify CLI option selection", 
         "Test if CLI correctly processes user input options", 
         "Application is ready to run", "choice = \"2\" (Test scraper only)", 
         "1. Run main()\n2. Enter \"2\" at the prompt\n3. Enter letter and limit", 
         "System correctly runs scraper test with specified parameters", "", ""],
        
        ["TC-CLI-002", "CLI", "User Interface", "Verify CLI with invalid input", 
         "Test system behavior with invalid CLI input", 
         "Application is ready to run", "choice = \"9\" (Invalid option)", 
         "1. Run main()\n2. Enter \"9\" at the prompt", 
         "System displays \"Invalid choice\" message and prompts again", "", ""],
        
        # Performance Tests
        ["TC-PER-001", "Performance", "Processing Speed", "Verify scraping performance", 
         "Measure time taken to scrape multiple articles", 
         "Internet connection available", "letter = \"A\"\nlimit = 10", 
         "1. Measure time to run test_scraper(\"A\", 10)", 
         "Scraping completes within acceptable time limit (< 2 minutes for 10 articles)", "", ""],
        
        ["TC-PER-002", "Performance", "Memory Usage", "Verify memory usage during vectorization", 
         "Monitor memory consumption during vectorization of large dataset", 
         "Large dataset available (>20 articles)", "input_dir with >20 articles", 
         "1. Monitor memory usage during vectorizer.process()", 
         "Memory usage remains within acceptable limits", "", ""],
        
        # Security Tests
        ["TC-SEC-001", "Security", "Data Handling", "Verify safe URL handling", 
         "Test if system safely handles potentially malicious URLs", 
         "N/A", "url with SQL injection attempt", 
         "1. Call scraper.retrieve_webpage() with potentially malicious URL", 
         "System handles URL safely without security issues", "", ""],
    ]
    
    # Define column headers
    columns = [
        "TEST CASE ID", "SECTION", "SUB-SECTION", "TEST CASE TITLE", "TEST DESCRIPTION", 
        "PRECONDITIONS", "TEST DATA", "TEST STEPS", "EXPECTED RESULT", "ACTUAL RESULT", "STATUS"
    ]
    
    # Create DataFrame
    df = pd.DataFrame(test_cases, columns=columns)
    
    # Create Excel writer
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    
    # Write DataFrame to Excel
    df.to_excel(writer, sheet_name='Test Cases', index=False)
    
    # Get the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Test Cases']
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Apply formatting to header row
    for col_num, column in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        
    # Set column widths based on content
    column_widths = {
        "TEST CASE ID": 12,
        "SECTION": 15,
        "SUB-SECTION": 15,
        "TEST CASE TITLE": 30,
        "TEST DESCRIPTION": 40,
        "PRECONDITIONS": 25,
        "TEST DATA": 25,
        "TEST STEPS": 40,
        "EXPECTED RESULT": 40,
        "ACTUAL RESULT": 20,
        "STATUS": 10
    }
    
    for i, column in enumerate(columns):
        worksheet.column_dimensions[get_column_letter(i+1)].width = column_widths.get(column, 15)
    
    # Apply borders and text wrapping to all cells
    for row in range(1, len(test_cases) + 2):  # +2 for header row and 1-indexing
        for col in range(1, len(columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Apply alternating row colors
    light_fill = PatternFill(start_color="EBF1F5", end_color="EBF1F5", fill_type="solid")
    for row in range(2, len(test_cases) + 2, 2):  # Every even row
        for col in range(1, len(columns) + 1):
            worksheet.cell(row=row, column=col).fill = light_fill
    
    # Group rows by section
    sections = {}
    for i, row in enumerate(test_cases, start=2):  # Start from row 2 (after header)
        section = row[1]
        if section not in sections:
            sections[section] = {"start": i, "end": i}
        else:
            sections[section]["end"] = i
    
    # Save the Excel file
    writer.close()
    
    print(f"Test cases successfully generated and saved to {output_file}")
    return output_file

if __name__ == "__main__":
    generate_test_cases_excel()